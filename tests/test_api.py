import time
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from openpyxl import load_workbook
def test_home_active_name(client):
    r=client.get("/"); assert r.status_code==200 and "Default Setup" in r.text
    assert f'user:' not in r.text and client.test_username in r.text
    assert "Last 90 days" in r.text and "Custom range" in r.text
    assert "Publication Date &amp; Time" in r.text
    js=client.get("/static/app.js"); assert "function renderImportPreview()" in js.text
    assert "function formatPublicationDate(value)" in js.text and 'timeZoneName:"short"' in js.text
    assert "active-pill" in js.text and client.get("/static/setups.css").status_code==200
    assert "function setScanVisual" in js.text and client.get("/static/scan.css").status_code==200
    assert 'id="nextAutomaticScan"' in r.text and "function startAutomaticScanCountdown" in js.text
    assert "Automatic scan in ${minutes} min" in js.text
    assert "shield-shape" in r.text and "click Scan Now" in r.text
    assert "Results update when you click Scan Now" in r.text and "window.scanIntervalSeconds" not in r.text
    assert "function updateAutoScanCountdown()" not in js.text and "await startScan()" not in js.text
    assert 'id="zeroDayScanBtn"' in r.text and "Scan for Zero Days" in r.text
    assert "function startZeroDayScan()" in js.text and "/api/scans/zero-days" in js.text
    assert 'id="zeroDayResults"' in r.text and "/api/zero-day-findings" in js.text
    assert "function zeroDayRowHtml(f)" in js.text
    assert 'id="confirmedZeroDayRows"' in r.text and 'id="activeExploitationRows"' in r.text
    assert "Confirmed Zero-Day Findings" in r.text and "Active Exploitation &amp; Other Priority Signals" in r.text
    assert client.get("/static/zero-day.css").status_code==200
    assert "Informational" in r.text and ">Unknown<" not in r.text
    assert "Scan every 30 minutes" in r.text and "9:00 a.m. Beirut time" in r.text
    assert "/static/my-threatlens-logo.png" in r.text
    assert client.get("/static/my-threatlens-logo.png").status_code==200
    assert 'id="appSidebar"' in r.text and "function toggleSidebar()" in js.text
    sidebar=r.text.split('<aside id="appSidebar"',1)[1].split('</aside>',1)[0]
    assert "Automatic Email" not in sidebar
    assert "✉ Automatic Email" in js.text and "openAutomaticEmail(${s.id}" in js.text
    assert 'id="autosaveLabel"' not in r.text and "Saved automatically" not in r.text
    assert "function scheduleAutoSave()" not in js.text and "function selectorDraftChanged()" in js.text
    assert 'onchange="selectorDraftChanged()"' in js.text and "if(!await saveSetup())return" in js.text
    assert "Your choices will be used when you click Scan Now." in r.text
    assert client.get("/static/sidebar.css").status_code==200
    assert client.get("/static/review.css").status_code==200
    assert "Checklist" not in js.text and "function saveChecklist(id)" not in js.text
    assert "function detailCves(cves)" in js.text and "Show ${links.length-8} more CVEs" in js.text
    assert "function renderFindingChat(messages)" in js.text and "async function loadFindingChat(id)" in js.text
    site_chat_js=client.get("/static/site-chat.js")
    assert 'id="siteChatLauncher"' in r.text and "function toggleSiteChat(open)" in site_chat_js.text
    assert client.get("/static/site-chat.css").status_code==200
    assert 'value==="undefined"' in js.text
    assert 'id="pagination" class="pagination" hidden' in r.text
    assert '$("#pagination").hidden=j.pages<=1' in js.text
    assert 'id="biggestThreat"' in r.text and "function renderBiggestThreat(f)" in js.text
    assert "Email this threat" in js.text and "openSingleEmail" in js.text
    assert 'id="emailSelected"' in r.text and "openSelectedEmail" in js.text
    assert 'id="emailAll"' in r.text and "openAllEmail" in js.text
    assert 'class="finding-select"' in js.text and "finding_ids" in js.text
    assert 'id="emailSubjectPreset"' in r.text and 'id="emailMessagePreset"' in r.text
    assert "Kindly check it now" in r.text and "applyEmailMessagePreset" in js.text
    assert client.get("/static/finding-email.css").status_code==200
    assert client.get("/static/biggest-threat.css").status_code==200
    about=client.get("/about"); assert about.status_code==200 and 'id="appSidebar"' in about.text
    assert 'id="siteChatLauncher"' in about.text and '/static/site-chat.js' in about.text
    assert "/?open=setups" in about.text and "/?open=import" in about.text
    assert "How to use the website" in about.text and "Define your scope" in about.text
    assert "AI-assisted summaries" in about.text and "independent workspace" in about.text
    assert "Why My ThreatLens?" in about.text and "Traditional vulnerability-management platforms" in about.text
    assert "tenable" not in about.text.lower()
    assert "does not collect user email addresses" in about.text and "one-way hash" in about.text
    assert 'href="/static/My-ThreatLens-Presentation.pptx"' in about.text
    assert "Download PowerPoint" in about.text
    assert client.get("/static/My-ThreatLens-Presentation.pptx").status_code==200
    assert client.get("/static/about.css").status_code==200
    assert "Arsen Eliane" in about.text and "ChatGPT" in about.text and "internship project" in about.text
    assert "Alfa" not in about.text and "Information Security Department" not in about.text
    assert 'new URLSearchParams(location.search).get("open")' in js.text
    assert 'link.download=`My-ThreatLens-Results-${stamp}.xlsx`' in js.text
    assert "scopeDirty=true" in js.text and "showPendingScopeState()" in js.text
    assert "Your search choices changed. Click Scan Now to get updated results." in js.text
    assert "Scanning selected sources for fresh findings" in js.text
    assert "Your additions" in js.text and "function addCustomOption()" in js.text and "function removeCustomOption(value)" in js.text
    assert 'class="manual-option-chip"' in js.text and client.get("/static/selector-custom.css").status_code==200

def test_shared_login_and_logout(client):
    from fastapi.testclient import TestClient
    with TestClient(client.app) as visitor:
        denied=visitor.get("/",follow_redirects=False)
        assert denied.status_code==303 and denied.headers["location"].startswith("/login")
        assert visitor.get("/api/setups").status_code==401
        assert visitor.get("/healthz").status_code==200
        login=visitor.get("/login")
        assert login.status_code==200 and "Every security headline" in login.text
        assert "Login ends when the browser closes" in login.text
        assert "Register" not in login.text
        assert visitor.get("/register",follow_redirects=False).status_code==303
        wrong=visitor.post("/login",data={"username":"cyber expert","password":"wrong"})
        assert wrong.status_code==401 and "incorrect" in wrong.text
        correct=visitor.post("/login",data={"username":"CYBER EXPERT","password":"test-only-password"},follow_redirects=False)
        assert correct.status_code==303 and visitor.cookies.get("mythreatlens_session")
        session_cookie=correct.headers["set-cookie"]
        assert "Max-Age" not in session_cookie and "Expires" not in session_cookie
        allowed=visitor.get("/")
        assert allowed.status_code==200 and "cyber expert" in allowed.text and allowed.headers["x-frame-options"]=="DENY"
        workspace_cookie=allowed.headers["set-cookie"]
        assert "threatlens_client=" in workspace_cookie and "Max-Age=31536000" in workspace_cookie
        logged_out=visitor.post("/logout",follow_redirects=False)
        assert logged_out.status_code==303 and logged_out.headers["location"]=="/login"
        assert visitor.get("/",follow_redirects=False).status_code==303

def test_browser_workspaces_are_isolated_and_locally_backed_up(client):
    from fastapi.testclient import TestClient
    with TestClient(client.app) as second:
        logged_in=second.post("/login",data={"username":"cyber expert","password":"test-only-password"},follow_redirects=False)
        assert logged_in.status_code==303
        assert client.get("/").status_code==200 and second.get("/").status_code==200
        created=client.post("/api/setups",json={"name":"Laptop Only","technologies":["Windows 11"],"keywords":["CVE"],"sources":["CISA"],"date_range":"7d"})
        assert created.status_code==201
        assert "Laptop Only" in [setup["name"] for setup in client.get("/api/setups").json()]
        assert "Laptop Only" not in [setup["name"] for setup in second.get("/api/setups").json()]
        assert client.cookies.get("mythreatlens_session")!=second.cookies.get("mythreatlens_session")
        workspace=client.get("/api/workspace").json()
        restored=second.post("/api/workspace/restore",json={"setups":[{"name":"Recovered Setup","technologies":["Windows 11"],"keywords":["CVE"],"sources":["CISA"],"date_range":"7d"}],"active_name":"Recovered Setup"})
        assert restored.status_code==200 and restored.json()["active"]["name"]=="Recovered Setup"
        assert workspace["instance_id"]==restored.json()["instance_id"]
    js=client.get("/static/app.js").text
    assert "localStorage.setItem(WORKSPACE_CACHE_KEY" in js and 'fetch("/api/workspace/restore"' in js
    assert "my-threatlens-browser-workspace-v1" in js
def test_custom_date_range_persists(client):
    data={"name":"Custom Dates","technologies":[],"keywords":[],"sources":[],"date_range":"custom","start_date":"2026-06-01","end_date":"2026-07-30"}
    created=client.post("/api/setups",json=data); assert created.status_code==201
    saved=created.json(); assert saved["date_range"]=="custom" and saved["start_date"]=="2026-06-01" and saved["end_date"]=="2026-07-30"
def test_setup_lifecycle(client):
    data={"name":"Ops","technologies":["FortiGate"],"keywords":["Exploit"],"sources":["CISA"],"date_range":"7d"}
    s=client.post("/api/setups",json=data); assert s.status_code==201
    sid=s.json()["id"]; data["name"]="Ops Renamed"; assert client.put(f"/api/setups/{sid}",json=data).json()["name"]=="Ops Renamed"
    assert client.post(f"/api/setups/{sid}/duplicate").status_code==200
    assert client.post(f"/api/setups/{sid}/activate").json()["active"]
    assert client.delete(f"/api/setups/{sid}").status_code==200
def test_import_preview_validation(client):
    from app.services.imports.service import sample_xlsx
    r=client.post("/api/import/preview",files={"file":("setup.xlsx",sample_xlsx(),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code==200 and r.json()["Setup Name"]=="Sample Setup"
    assert client.post("/api/import/preview",files={"file":("bad.txt",b"x","text/plain")}).status_code==400
def test_email_requires_valid_address_and_configuration(client,monkeypatch):
    invalid=client.post("/api/email",json={"recipient":"not-an-email","subject":"Report","message":"Attached"})
    assert invalid.status_code==422 and "valid recipient" in invalid.json()["detail"]
    from app.config import settings
    monkeypatch.setattr(settings,"smtp_password","")
    unconfigured=client.post("/api/email",json={"recipient":"analyst@example.com","subject":"Report","message":"Attached"})
    assert unconfigured.status_code==503 and "SMTP_PASSWORD" in unconfigured.json()["detail"]
def test_automatic_critical_email_is_deduplicated(monkeypatch):
    import asyncio
    from types import SimpleNamespace
    from app import main
    sent=[]
    monkeypatch.setattr(main.settings,"automatic_email_recipient","supervisor@example.com")
    monkeypatch.setattr(main.settings,"critical_email_enabled",True)
    monkeypatch.setattr(main,"send_findings_email",lambda settings,recipient,subject,body,rows,setup:sent.append([row.fingerprint for row in rows]))
    main.AUTOMATICALLY_ALERTED_FINGERPRINTS.clear()
    finding=SimpleNamespace(severity="Critical",fingerprint="critical-one",kev=False,ai_score=90,cvss=9.8,publication_date=datetime.now(timezone.utc))
    setup=SimpleNamespace(name="Default Setup")
    asyncio.run(main.send_new_critical_alerts([finding],setup))
    asyncio.run(main.send_new_critical_alerts([finding],setup))
    assert sent==[["critical-one"]]
def test_automatic_email_settings_are_saved_per_setup(client):
    initial=client.get("/api/automatic-email")
    assert initial.status_code==200 and initial.json()["recipients"]==[]
    saved=client.put("/api/automatic-email",json={"recipients":["IT@example.com","infosec@example.com","it@example.com"],"daily_enabled":True,"critical_enabled":False,"subject":"Security review","message":"Kindly check it now."})
    assert saved.status_code==200
    assert saved.json()["recipients"]==["it@example.com","infosec@example.com"]
    assert saved.json()["subject"]=="Security review"
    assert saved.json()["critical_enabled"] is False
    too_many=client.put("/api/automatic-email",json={"recipients":[f"user{i}@example.com" for i in range(11)]})
    assert too_many.status_code==422

def test_automatic_email_is_independent_for_each_setup(client):
    first=next(setup for setup in client.get("/api/setups").json() if setup["active"])
    second=client.post("/api/setups",json={"name":"Separate Team","technologies":["Windows 11"],"keywords":["CVE"],"sources":["CISA"],"date_range":"7d"}).json()
    saved_first=client.put(f'/api/automatic-email?setup_id={first["id"]}',json={"recipients":["first@example.com"],"subject":"First setup","message":"First team message."})
    saved_second=client.put(f'/api/automatic-email?setup_id={second["id"]}',json={"recipients":["second@example.com","backup@example.com"],"subject":"Second setup","message":"Second team message."})
    assert saved_first.status_code==saved_second.status_code==200
    first_config=client.get(f'/api/automatic-email?setup_id={first["id"]}').json()
    second_config=client.get(f'/api/automatic-email?setup_id={second["id"]}').json()
    assert first_config["recipients"]==["first@example.com"] and first_config["setup_name"]==first["name"]
    assert second_config["recipients"]==["second@example.com","backup@example.com"] and second_config["setup_name"]==second["name"]
    assert client.get("/api/automatic-email?setup_id=999999").status_code==404

def test_automatic_scheduler_has_daily_and_30_minute_jobs(client):
    from app import main
    assert main.AUTOMATION_SCHEDULER.get_job("daily-email")
    job=main.AUTOMATION_SCHEDULER.get_job("critical-scan")
    assert job and int(job.trigger.interval.total_seconds())==1800

def test_automatic_30_minute_scan_runs_configured_setup(client,monkeypatch):
    import asyncio
    from sqlalchemy import delete
    from app import main
    from app.database import SessionLocal
    from app.models import EmailAutomation
    with SessionLocal() as db:
        db.execute(delete(EmailAutomation)); db.commit()
    client.put("/api/automatic-email",json={"recipients":["supervisor@example.com"],"daily_enabled":True,"critical_enabled":True})
    started=[]
    async def fake_run_scan(scan_id): started.append(main.SCANS_CACHE[scan_id])
    monkeypatch.setattr(main,"run_scan",fake_run_scan)
    main.SCANS_CACHE.clear()
    asyncio.run(main.run_automatic_critical_scans())
    assert len(started)==1 and started[0]["message"]=="Automatic 30-minute scan queued"

def test_daily_email_contains_only_critical_findings(client,monkeypatch):
    import asyncio
    from types import SimpleNamespace
    from sqlalchemy import delete
    from app import main
    from app.database import SessionLocal
    from app.models import EmailAutomation
    with SessionLocal() as db:
        db.execute(delete(EmailAutomation)); db.commit()
    client.put("/api/automatic-email",json={"recipients":["supervisor@example.com"],"daily_enabled":True,"critical_enabled":True})
    monkeypatch.setattr(main,"run_scan",lambda scan_id: asyncio.sleep(0))
    findings=[SimpleNamespace(severity="Critical",kev=False,ai_score=90,cvss=9.8,publication_date=datetime.now(timezone.utc)),SimpleNamespace(severity="High",kev=False,ai_score=80,cvss=8.0,publication_date=datetime.now(timezone.utc))]
    monkeypatch.setattr(main,"scoped_findings",lambda db,setup,params:findings)
    delivered=[]
    async def fake_delivery(subject,body,rows,setup,status_key,recipients): delivered.extend(rows); return True
    monkeypatch.setattr(main,"deliver_automatic_email",fake_delivery)
    asyncio.run(main.run_daily_email_report())
    assert [finding.severity for finding in delivered]==["Critical"]
def test_automatic_email_test_sends_to_every_saved_recipient(client,monkeypatch):
    from app import main
    sent=[]
    client.put("/api/automatic-email",json={"recipients":["it@example.com","infosec@example.com"]})
    monkeypatch.setattr(main,"send_findings_email",lambda settings,recipient,subject,body,rows,setup:sent.append(recipient))
    result=client.post("/api/automatic-email/test")
    assert result.status_code==200 and result.json()["sent_count"]==2
    assert sent==["it@example.com","infosec@example.com"]
def test_manual_critical_automation_check_uses_saved_setup(client,monkeypatch):
    from app import main
    assert client.post("/api/automatic-email/run-now").status_code==422
    client.put("/api/automatic-email",json={"recipients":["supervisor@example.com"],"critical_enabled":True})
    completed=[]
    async def fake_run_scan(scan_id):
        completed.append(scan_id)
        main.SCANS_CACHE[scan_id].update(status="completed",progress=100,automation={"email_sent":False,"critical_total":0,"new_critical_count":0,"recipient_count":1})
    monkeypatch.setattr(main,"run_scan",fake_run_scan)
    result=client.post("/api/automatic-email/run-now")
    assert result.status_code==202 and completed==[result.json()["scan_id"]]
    status=client.get(f'/api/scans/{result.json()["scan_id"]}').json()
    assert status["status"]=="completed" and status["automation"]["recipient_count"]==1
def test_scan_202_completion_filters_export(client,monkeypatch):
    setups=client.get("/api/setups").json(); default=next(s for s in setups if s["name"]=="Default Setup"); client.post(f'/api/setups/{default["id"]}/activate')
    r=client.post("/api/scans"); assert r.status_code==202
    sid=r.json()["scan_id"]
    for _ in range(30):
        status=client.get(f"/api/scans/{sid}").json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert status["status"]=="completed"
    assert {"matched_total","in_range","excluded_by_date","recommended_range","sources_checked","live_sources","unavailable_sources"}.issubset(status["metrics"])
    result=client.get("/api/findings?severity=Critical").json(); assert result["page"]==1 and result["pages"]>=1
    assert result["biggest"] and result["biggest"]["id"]
    sent={}
    monkeypatch.setattr("app.main.send_findings_email",lambda settings,recipient,subject,message,rows,setup:sent.update(rows=rows,subject=subject))
    emailed=client.post(f'/api/email?finding_id={result["biggest"]["id"]}',json={"recipient":"supervisor@example.com","subject":"Biggest threat","message":"Review this threat."})
    assert emailed.status_code==200 and emailed.json()["findings_count"]==1
    assert len(sent["rows"])==1 and sent["rows"][0].id==result["biggest"]["id"]
    available=client.get("/api/findings").json()["items"]
    selected_ids=[finding["id"] for finding in available[:2]]
    assert len(selected_ids)==2
    selected=client.post("/api/email",json={"recipient":"supervisor@example.com","subject":"Selected threats","message":"Review these threats.","finding_ids":selected_ids})
    assert selected.status_code==200 and selected.json()["findings_count"]==2
    assert [finding.id for finding in sent["rows"]]==selected_ids
    all_critical=client.post("/api/email?severity=Critical",json={"recipient":"supervisor@example.com","subject":"All critical threats","message":"Kindly check it now."})
    assert all_critical.status_code==200 and all_critical.json()["findings_count"]==result["total"]
    assert len(sent["rows"])==result["total"] and all(finding.severity=="Critical" for finding in sent["rows"])
    if result["items"]: assert result["items"][0]["ai_summary"] and result["items"][0]["ai_reason"]
    exp=client.get("/api/export?severity=Critical"); wb=load_workbook(BytesIO(exp.content))
    import re
    assert re.search(r'My-ThreatLens-Results-\d{8}-\d{6}\.xlsx',exp.headers["content-disposition"])
    assert set(["Results","Export Context","Finding Details","Additional CVEs"])<=set(wb.sheetnames)
    assert "Review Checklist" not in wb.sheetnames and "Review Progress" not in [cell.value for cell in wb["Results"][1]]
    assert wb["Results"].max_row-1==result["total"]

def test_scan_count_uses_selected_date_range(client):
    data={"name":"Recent Palo Alto","technologies":["Palo Alto Networks"],"keywords":["Exploit"],"sources":["The Hacker News"],"date_range":"1d"}
    setup=client.post("/api/setups",json=data).json()
    scan=client.post("/api/scans").json()
    for _ in range(30):
        status=client.get(f'/api/scans/{scan["scan_id"]}').json()
        if status["status"]=="completed": break
        time.sleep(.03)
    visible=client.get("/api/findings").json()["total"]
    assert status["findings_count"]==visible==0
    assert "no findings were published in the selected date range" in status["message"]
    assert status["metrics"]["excluded_by_date"]==2

def test_selecting_entire_catalog_does_not_filter_out_unlisted_story_terms(client):
    from app.services.matching.aliases import TECH_ALIASES, KEYWORD_ALIASES
    client.post("/api/setups",json={"name":"Full Catalog","technologies":list(TECH_ALIASES),"keywords":list(KEYWORD_ALIASES),"sources":["The Hacker News"],"date_range":"7d"})
    scan=client.post("/api/scans").json()
    for _ in range(30):
        status=client.get(f'/api/scans/{scan["scan_id"]}').json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert status["findings_count"]==12
    assert client.get("/api/findings").json()["total"]==12

def test_zero_day_addon_scans_independently_of_regular_keywords(client):
    client.post("/api/setups",json={"name":"Zero-Day Watch","technologies":["Windows 11","Exchange Server"],"keywords":["SQL Injection"],"sources":["The Hacker News"],"date_range":"7d"})
    scan=client.post("/api/scans/zero-days")
    assert scan.status_code==202
    scan_id=scan.json()["scan_id"]
    for _ in range(30):
        status=client.get(f"/api/scans/{scan_id}").json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert status["status"]=="completed" and status["kind"]=="zero_day"
    assert status["findings_count"]>=2
    assert status["metrics"]["zero_day_mentions"]>=1
    assert status["metrics"]["active_exploitation"]>=1
    assert status["metrics"]["critical_priority"]>=1
    assert status["metrics"]["sources_checked"]==1
    normal_findings=client.get("/api/findings").json()
    zero_day_findings=client.get("/api/zero-day-findings").json()
    assert normal_findings["total"]==0
    assert zero_day_findings["scanned"] is True and zero_day_findings["total"]==status["findings_count"]
    assert zero_day_findings["total"]==zero_day_findings["zero_day_total"]+zero_day_findings["active_exploitation_total"]
    assert zero_day_findings["zero_day_total"]>=1 and zero_day_findings["active_exploitation_total"]>=1
    assert all(finding["priority_category"]=="zero_day" for finding in zero_day_findings["zero_days"])
    assert all(finding["priority_category"]=="active_exploitation" for finding in zero_day_findings["active_exploitation"])
    assert zero_day_findings["metrics"]==status["metrics"] and len(zero_day_findings["sources"])==1
    assert all("SQL Injection" not in finding["matched_keywords"] for finding in zero_day_findings["items"])
    normal_scan=client.post("/api/scans").json()
    for _ in range(30):
        normal_status=client.get(f'/api/scans/{normal_scan["scan_id"]}').json()
        if normal_status["status"]=="completed": break
        time.sleep(.03)
    assert client.get("/api/findings").json()["total"]==0
    assert client.get("/api/zero-day-findings").json()["total"]==zero_day_findings["total"]

def test_zero_day_scan_completes_safely_when_a_collector_crashes(client,monkeypatch):
    from app import main
    async def broken_collector(*args,**kwargs): raise RuntimeError("network failure")
    monkeypatch.setattr(main,"collect_source",broken_collector)
    client.post("/api/setups",json={"name":"Resilient Zero-Day","technologies":["Windows 11"],"keywords":["CVE"],"sources":["CISA"],"date_range":"7d"})
    queued=client.post("/api/scans/zero-days").json()
    for _ in range(30):
        status=client.get(f'/api/scans/{queued["scan_id"]}').json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert status["status"]=="completed" and status["findings_count"]==0
    assert status["metrics"]["sources_checked"]==1 and status["metrics"]["live_sources"]==0
    assert status["sources"][0]["status"]=="unavailable"
    results=client.get("/api/zero-day-findings").json()
    assert results["scanned"] is True and results["total"]==0
def test_delete_setup_with_scan_history(client):
    data={"name":"Delete With History","technologies":["FortiGate"],"keywords":["Exploit"],"sources":["CISA"],"date_range":"7d"}
    setup=client.post("/api/setups",json=data).json()
    scan=client.post("/api/scans"); assert scan.status_code==202
    sid=scan.json()["scan_id"]
    for _ in range(30):
        status=client.get(f"/api/scans/{sid}").json()
        if status["status"]=="completed": break
        time.sleep(.03)
    deleted=client.delete(f'/api/setups/{setup["id"]}')
    assert deleted.status_code==200 and deleted.json()["deleted"] is True
    assert all(s["id"]!=setup["id"] for s in client.get("/api/setups").json())
def test_scan_results_are_not_persisted_to_sqlite(client):
    from sqlalchemy import func,select
    from app.database import SessionLocal
    from app.models import Finding,Scan,ChatMessage,SourceStatus
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(Finding))==0
        assert db.scalar(select(func.count()).select_from(Scan))==0
        assert db.scalar(select(func.count()).select_from(ChatMessage))==0
        assert db.scalar(select(func.count()).select_from(SourceStatus))==0
def test_last_setup_cannot_be_deleted(client):
    setups=client.get("/api/setups").json()
    for setup in setups[1:]: client.delete(f'/api/setups/{setup["id"]}')
    remaining=client.get("/api/setups").json()
    assert len(remaining)==1
    response=client.delete(f'/api/setups/{remaining[0]["id"]}')
    assert response.status_code==409 and "last setup cannot be deleted" in response.json()["detail"]
def test_launcher_replaces_only_older_threatlens_server():
    launcher=open("START_MY_THREATLENS.bat",encoding="utf-8").read()
    assert "Get-NetTCPConnection -LocalPort 8001" in launcher
    assert "uvicorn.+app\\.main:app" in launcher
    assert "Port 8001 is used by another application" in launcher
def test_results_follow_active_setup_source_and_date_range(client):
    setups=client.get("/api/setups").json()
    target=setups[0]
    client.post(f'/api/setups/{target["id"]}/activate')
    payload={"name":target["name"],"description":"","technologies":["Exchange Server"],"keywords":["CVE"],"sources":["The Hacker News"],"date_range":"1d","start_date":None,"end_date":None}
    assert client.put(f'/api/setups/{target["id"]}',json=payload).status_code==200
    result=client.get("/api/findings").json()
    assert all(item["source"]=="The Hacker News" and item["technology"]=="Exchange Server" for item in result["items"])
def test_results_refresh_when_setup_keywords_change(client):
    setup=client.post("/api/setups",json={"name":"Keyword Refresh","technologies":["Palo Alto Networks"],"keywords":["Exploit"],"sources":["The Hacker News"],"date_range":"7d"}).json()
    scan=client.post("/api/scans").json()
    for _ in range(30):
        status=client.get(f'/api/scans/{scan["scan_id"]}').json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert client.get("/api/findings").json()["total"]>0
    changed={"name":setup["name"],"description":"","technologies":["Palo Alto Networks"],"keywords":["SQL Injection"],"sources":["The Hacker News"],"date_range":"7d","start_date":None,"end_date":None}
    assert client.put(f'/api/setups/{setup["id"]}',json=changed).status_code==200
    assert client.get("/api/findings").json()["total"]==0
    changed["keywords"]=["RCE"]
    assert client.put(f'/api/setups/{setup["id"]}',json=changed).status_code==200
    rescan=client.post("/api/scans").json()
    for _ in range(30):
        status=client.get(f'/api/scans/{rescan["scan_id"]}').json()
        if status["status"]=="completed": break
        time.sleep(.03)
    assert client.get("/api/findings").json()["total"]>0
def test_custom_dates_hidden_by_default(client):
    css=client.get("/static/date-range.css").text
    assert ".customdates[hidden]" in css and "display: none" in css
    sidebar_css=client.get("/static/sidebar.css").text
    assert ".pagination[hidden]" in sidebar_css
def test_chat_grounded(client,monkeypatch):
    async def fake_ollama(finding,history,settings): return "Grounded local-model response"
    monkeypatch.setattr("app.main.ollama_answer",fake_ollama)
    findings=client.get("/api/findings").json()["items"]
    if findings:
        a=client.post(f'/api/findings/{findings[0]["id"]}/chat',json={"question":"Are we affected?"}).json()
        assert a["message"]["content"]=="Grounded local-model response"
        history=client.get(f'/api/findings/{findings[0]["id"]}/chat').json()["messages"]
        assert [message["role"] for message in history[-2:]]==["user","assistant"]

def test_site_chat_uses_workspace_context(client,monkeypatch):
    captured={}
    async def fake_site_ollama(setup,findings,history,settings):
        captured.update(setup=setup.display_name,findings=len(findings),question=history[-1]["content"],model=settings.ollama_model)
        return "Workspace-level Ollama response"
    monkeypatch.setattr("app.main.ollama_site_answer",fake_site_ollama)
    response=client.post("/api/site-chat",json={"question":"What should I prioritize?"})
    assert response.status_code==200 and response.json()["message"]["content"]=="Workspace-level Ollama response"
    assert captured["setup"] and captured["question"]=="What should I prioritize?"
    assert captured["model"]=="gpt-oss:20b"
    history=client.get("/api/site-chat").json()["messages"]
    assert [message["role"] for message in history]==["user","assistant"]
    assert client.delete("/api/site-chat").json()["deleted"] is True
    assert client.get("/api/site-chat").json()["messages"]==[]

def test_site_chat_rejects_empty_or_oversized_questions(client):
    assert client.post("/api/site-chat",json={"question":"   "}).status_code==422
    assert client.post("/api/site-chat",json={"question":"x"*2001}).status_code==422

def test_chat_does_not_claim_unrelated_platform_is_affected():
    from app.services.chat import answer
    finding=SimpleNamespace(title="Cloud credential exposure",summary="Leaked credentials were used against AWS services.",technology="AWS",matched_technologies=["AWS"],matched_keywords=["Credential Stuffing"],source="Example",publication_date=datetime.now(timezone.utc),severity="High",severity_basis="Source assessment",cves=[],url="https://example.test/finding")
    result=answer(finding,"How am I affected on Windows 11?")
    assert "No direct impact on Windows 11" in result["direct_answer"]
    assert "matches AWS, not Windows 11" in result["direct_answer"]
    assert result["engine"]=="deterministic-grounded-v3"

def test_chat_answers_about_a_matched_platform():
    from app.services.chat import answer
    finding=SimpleNamespace(title="Windows vulnerability",summary="A Windows Shell flaw allows privilege escalation.",technology="Windows 11",matched_technologies=["Windows 11"],matched_keywords=["Privilege Escalation"],source="Example",publication_date=datetime.now(timezone.utc),severity="High",severity_basis="CVSS",cves=["CVE-2026-12345"],url="https://example.test/finding")
    result=answer(finding,"Am I vulnerable on Windows 11?")
    assert "could be affected if you operate Windows 11" in result["direct_answer"]
    assert any("CVE-2026-12345" in fact for fact in result["verified_facts"])

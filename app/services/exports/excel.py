from io import BytesIO
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def safe(v):
    s="" if v is None else str(v)
    return "'"+s if s.startswith(("=","+","-","@")) else s
def create_workbook(findings, setup, filters):
    wb=Workbook(); ws=wb.active; ws.title="Results"
    heads=["Severity","Technology","Finding","Summary","Publication Date","CVEs (max 5)","Source","Source URL","AI Relevance","AI Confidence"]
    ws.append(heads); ws.freeze_panes="A2"; ws.auto_filter.ref="A1:J1"
    colors={"Critical":"D92D20","High":"F79009","Medium":"FEC84B","Low":"12B76A","Informational":"667085"}
    for i,f in enumerate(findings,2):
        ws.append([f.severity,f.technology,safe(f.title),safe(f.summary),f.publication_date.replace(tzinfo=None),", ".join(f.cves[:5]),f.source,f.url,f"{f.ai_score}/100",f.ai_confidence])
        ws.cell(i,1).fill=PatternFill("solid",fgColor=colors.get(f.severity,"667085")); ws.cell(i,1).font=Font(color="FFFFFF",bold=True)
        ws.cell(i,3).hyperlink=f.url; ws.cell(i,8).hyperlink=f.url
        if i%2==0:
            for c in range(2,11): ws.cell(i,c).fill=PatternFill("solid",fgColor="F4F8FB")
    widths=[12,22,45,55,20,34,24,45,15,15]
    for n,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(n)].width=w
    for row in ws.iter_rows():
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
    for cell in ws[1]: cell.fill=PatternFill("solid",fgColor="075985"); cell.font=Font(color="FFFFFF",bold=True)
    ctx=wb.create_sheet("Export Context")
    setup_name=getattr(setup,"display_name",setup.name)
    rows=[("Exported At",datetime.now(timezone.utc).isoformat()),("Active Setup Name",setup_name),("Selected Technologies",", ".join(setup.technologies)),("Selected Keywords",", ".join(setup.keywords)),("Selected Sources",", ".join(setup.sources)),("Date Range",setup.date_range),("Active Filters",str(filters)),("Total Exported Findings",len(findings))]
    for r in rows: ctx.append(r)
    details=wb.create_sheet("Finding Details"); details.append(["Finding","CVSS","EPSS","KEV","Severity basis","AI reason","Evidence links","Notes"])
    extra=wb.create_sheet("Additional CVEs"); extra.append(["Finding","Additional CVE"])
    for f in findings:
        details.append([safe(f.title),f.cvss,f.epss,"Yes" if f.kev else "No",f.severity_basis,f.ai_reason,"\n".join(f.evidence),safe(f.notes)])
        for c in f.cves[5:]: extra.append([safe(f.title),c])
    for sheet in wb:
        sheet.freeze_panes="A2"
        for cell in sheet[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="075985")
    out=BytesIO(); wb.save(out); return out.getvalue()
